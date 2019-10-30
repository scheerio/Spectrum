import openpyxl
import pyperclip
from tkinter import *

#LOAD AND ACTIVATE WORKBOOK
book = openpyxl.load_workbook(r'C:\Users\bscheer\Desktop\Key References\IA Elaboration Roster_20191029.xlsx')
ws = book.active

#ARRAYS HOLD FIRST AND LAST NAMES
firstarray=[]
lastarray=[]

#STRINGS TO CREATE CODE TEXT
firsttext='firstname=['
lasttext='lastname=['
emailtext='email=['
focustext='focus=['
nicktext='nickname=['

#SPLIT UP THE FULL NAMES INTO FIRST AND LAST
firstone=True
for cell in ws['A']:
	if (cell.value==None):
		continue
	if (firstone==False):	
		temp=[]
		temp=cell.value.split(' ')
		firstarray.append(temp[0])
		lastarray.append(temp[1])
	else:
		firstone=False

#TURN FIRST NAME LIST INTO CODE TEXT	
count=0
for x in firstarray:
	firsttext = firsttext + "\""+x+"\""
	count=count+1
	if (count==len(firstarray)):
		firsttext = firsttext + ']'
	else:
		firsttext=firsttext + ', '

#TURN LAST NAME LIST INTO CODE TEXT
count2=0
for x in lastarray:
	lasttext = lasttext + "\""+x+"\""
	count2=count2+1
	if (count2==len(lastarray)):
		lasttext = lasttext + ']'
	else:
		lasttext=lasttext + ', '

#TURN EMAIL LIST INTO CODE TEXT
skip=True
firstone=True
for cell in ws['D']:
	if (cell.value==None):
		continue
	if (skip==False):
		if (firstone==False):
			emailtext=emailtext + ', '
		emailtext=emailtext + "\""+cell.value+"\""
		firstone=False
	skip=False
emailtext=emailtext + ']'

#TURN ORGANIZATION LIST INTO CODE TEXT
skip2=True
firstone2=True
for cell in ws['C']:
	if (cell.value==None):
		continue
	if (skip2==False):
		if (firstone2==False):
			focustext=focustext + ', '
		focustext=focustext + "\""+cell.value+"\""
		firstone2=False
	skip2=False
focustext=focustext + ']'

#TURN NICKNAME LIST INTO CODE TEXT
skip3=True
firstone3=True
for cell in ws['B']:
	if (cell.value==None):
		continue
	if (skip3==False):
		if (firstone3==False):
			nicktext=nicktext + ', '
		nicktext=nicktext + "\""+cell.value+"\""
		firstone3=False
	skip3=False
nicktext=nicktext + ']'

#PUT TOGETHER FINAL PRODUCT AND COPY TO CLIPBOARD
print('\nSuccess! The below code is copied to your clipboard. Add it into spectrum.py!\n')
text=firsttext+'\n\n'+lasttext+'\n\n'+emailtext+'\n\n'+focustext+'\n\n'+nicktext
print(text)
pyperclip.copy(text)

#DISPLAY SUCCESS MESSAGE
root = Tk()
labelfont = ('times', 20, 'bold')   
root.title('Success Confirmation')               
widget = Label(root, text='The following was copied to your clipboard:\n\n'+text, wraplength=600, justify=LEFT)
widget.config(height=35, width=90)
widget.pack(expand=YES, fill=BOTH)
root.mainloop()

with open("translater.py", "w") as f1:
        f1.writelines(text)












